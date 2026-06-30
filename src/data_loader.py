import csv
import warnings
from io import BytesIO, StringIO
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


CANONICAL_COLUMNS = [
    "platform",
    "account_name",
    "store_name",
    "campaign_name",
    "ad_group_name",
    "ad_name",
    "ad_id",
    "product_code",
    "status",
    "bidding_mode",
    "placement",
    "date",
    "period_start",
    "period_end",
    "granularity",
    "impressions",
    "clicks",
    "ctr",
    "add_to_cart",
    "add_to_cart_rate",
    "conversions",
    "direct_conversions",
    "conversion_rate",
    "direct_conversion_rate",
    "cpa",
    "direct_cpa",
    "units_sold",
    "direct_units_sold",
    "revenue",
    "direct_revenue",
    "spend",
    "roas",
    "direct_roas",
    "acos",
    "direct_acos",
    "voucher_amount",
    "vouchered_sales",
    "source_file",
]


TEMPLATE_ALIASES = {
    "platform": "platform",
    "account": "account_name",
    "account_name": "account_name",
    "store": "store_name",
    "store_name": "store_name",
    "campaign": "campaign_name",
    "campaign_name": "campaign_name",
    "ad_group": "ad_group_name",
    "ad_group_name": "ad_group_name",
    "ad": "ad_name",
    "ad_name": "ad_name",
    "ad_id": "ad_id",
    "product_code": "product_code",
    "status": "status",
    "bidding_mode": "bidding_mode",
    "placement": "placement",
    "date": "date",
    "period_start": "period_start",
    "period_end": "period_end",
    "granularity": "granularity",
    "impressions": "impressions",
    "clicks": "clicks",
    "ctr": "ctr",
    "add_to_cart": "add_to_cart",
    "add_to_cart_rate": "add_to_cart_rate",
    "conversions": "conversions",
    "direct_conversions": "direct_conversions",
    "conversion_rate": "conversion_rate",
    "direct_conversion_rate": "direct_conversion_rate",
    "cpa": "cpa",
    "direct_cpa": "direct_cpa",
    "units_sold": "units_sold",
    "direct_units_sold": "direct_units_sold",
    "revenue": "revenue",
    "direct_revenue": "direct_revenue",
    "spend": "spend",
    "roas": "roas",
    "direct_roas": "direct_roas",
    "acos": "acos",
    "direct_acos": "direct_acos",
    "voucher_amount": "voucher_amount",
    "vouchered_sales": "vouchered_sales",
}


SHOPEE_MAP = {
    "Nama Iklan": "ad_name",
    "Status": "status",
    "Kode Produk": "product_code",
    "Mode Bidding": "bidding_mode",
    "Penempatan Iklan": "placement",
    "Dilihat": "impressions",
    "Jumlah Klik": "clicks",
    "Persentase Klik": "ctr",
    "Add to Cart": "add_to_cart",
    "Add to Cart Rate": "add_to_cart_rate",
    "Konversi": "conversions",
    "Konversi Langsung": "direct_conversions",
    "Tingkat konversi": "conversion_rate",
    "Tingkat Konversi Langsung": "direct_conversion_rate",
    "Biaya per Konversi": "cpa",
    "Biaya per Konversi Langsung": "direct_cpa",
    "Produk Terjual": "units_sold",
    "Terjual Langsung": "direct_units_sold",
    "Omzet Penjualan": "revenue",
    "Penjualan Langsung (GMV Langsung)": "direct_revenue",
    "Biaya": "spend",
    "Efektifitas Iklan": "roas",
    "Efektivitas Langsung": "direct_roas",
    "Persentase Biaya Iklan terhadap Penjualan dari Iklan (ACOS)": "acos",
    "Persentase Biaya Iklan terhadap Penjualan dari Iklan Langsung (ACOS Langsung)": "direct_acos",
    "Voucher Amount": "voucher_amount",
    "Vouchered Sales": "vouchered_sales",
}


SHOPEE_CAMPAIGN_TYPE_MAP = {
    "Shop Name": "store_name",
    "Campaign Type": "campaign_name",
    "Product Name": "ad_name",
    "Product ID": "product_code",
    "Impressions": "impressions",
    "Clicks": "clicks",
    "CTR": "ctr",
    "Ads Spend(Local currency)": "spend",
    "CPC": "cpc",
    "Orders": "conversions",
    "Gross Sales(Local currency)": "revenue",
    "ROAS": "roas",
    "Units Sold": "units_sold",
    "CR": "conversion_rate",
}


NUMERIC_COLUMNS = [
    "impressions",
    "clicks",
    "ctr",
    "add_to_cart",
    "add_to_cart_rate",
    "conversions",
    "direct_conversions",
    "conversion_rate",
    "direct_conversion_rate",
    "cpa",
    "direct_cpa",
    "units_sold",
    "direct_units_sold",
    "revenue",
    "direct_revenue",
    "spend",
    "roas",
    "direct_roas",
    "acos",
    "direct_acos",
    "voucher_amount",
    "vouchered_sales",
]


def _seek_start(file_obj):
    if hasattr(file_obj, "seek"):
        file_obj.seek(0)


def _clean_number(value):
    if pd.isna(value):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if text in {"", "-"}:
        return 0.0
    is_percent = text.endswith("%")
    text = text.replace("%", "").replace(",", "").replace("IDR", "").strip()
    try:
        number = float(text)
    except ValueError:
        return 0.0
    if is_percent:
        return number / 100
    return number


def _clean_identifier(value):
    if pd.isna(value):
        return pd.NA
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    text = str(value).strip()
    if text in {"", "-"}:
        return pd.NA
    return text


def _normalize_dates(df):
    for column in ["date", "period_start", "period_end"]:
        if column in df.columns:
            df[column] = pd.to_datetime(df[column], errors="coerce", dayfirst=True)
    return df


def _finalize(df, source_file):
    for column in CANONICAL_COLUMNS:
        if column not in df.columns:
            df[column] = pd.NA

    for column in NUMERIC_COLUMNS:
        df[column] = df[column].apply(_clean_number)

    df = _normalize_dates(df)
    df["source_file"] = source_file
    df["platform"] = df["platform"].fillna("Unknown")
    df["status"] = df["status"].fillna("Unknown")
    df["granularity"] = df["granularity"].fillna("period")
    for column in ["ad_id", "product_code"]:
        df[column] = df[column].apply(_clean_identifier)

    if df["roas"].eq(0).all():
        df["roas"] = df["revenue"] / df["spend"].replace(0, pd.NA)
        df["roas"] = pd.to_numeric(df["roas"], errors="coerce").fillna(0)
    if df["acos"].eq(0).all():
        df["acos"] = df["spend"] / df["revenue"].replace(0, pd.NA)
        df["acos"] = pd.to_numeric(df["acos"], errors="coerce").fillna(0)

    return df[CANONICAL_COLUMNS]


def _read_text(file_obj):
    if isinstance(file_obj, (str, Path)):
        return Path(file_obj).read_text(encoding="utf-8-sig")
    raw = file_obj.getvalue()
    if isinstance(raw, str):
        return raw
    return raw.decode("utf-8-sig")


def _metadata_value(lines, label):
    for line in lines:
        cells = next(csv.reader([line]))
        if cells and cells[0] == label and len(cells) > 1:
            return cells[1]
    return None


def _parse_period(text):
    if not text or " - " not in text:
        return pd.NaT, pd.NaT
    start, end = text.split(" - ", 1)
    return pd.to_datetime(start, dayfirst=True, errors="coerce"), pd.to_datetime(
        end, dayfirst=True, errors="coerce"
    )


def _is_shopee_export(text):
    return "Laporan Iklan Produk - Shopee Indonesia" in text and "Nama Iklan" in text


def _load_shopee_csv(file_obj, source_file):
    text = _read_text(file_obj)
    lines = text.splitlines()
    header_index = None
    for index, line in enumerate(lines):
        if line.startswith("Urutan,Nama Iklan,Status"):
            header_index = index
            break
    if header_index is None:
        raise ValueError("Could not find Shopee Ads table header.")

    table_text = "\n".join(lines[header_index:])
    df = pd.read_csv(StringIO(table_text))
    df = df.rename(columns=SHOPEE_MAP)

    period_start, period_end = _parse_period(_metadata_value(lines, "Periode"))
    df["platform"] = "Shopee Ads"
    df["account_name"] = _metadata_value(lines, "Username")
    df["store_name"] = _metadata_value(lines, "Nama Toko")
    df["campaign_name"] = df["ad_name"]
    df["period_start"] = period_start
    df["period_end"] = period_end
    df["granularity"] = "period"
    return _finalize(df, source_file)


def _load_template_dataframe(df, source_file):
    renamed = {}
    for column in df.columns:
        key = str(column).strip().lower().replace(" ", "_")
        if key in TEMPLATE_ALIASES:
            renamed[column] = TEMPLATE_ALIASES[key]
    normalized = df.rename(columns=renamed)
    return _finalize(normalized, source_file)


def _worksheet_to_dataframe(file_obj, sheet_name):
    _seek_start(file_obj)
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message="Workbook contains no default style")
        workbook = load_workbook(file_obj, read_only=False, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError("Workbook does not contain sheet '{}'.".format(sheet_name))

    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    rows = [row for row in rows if any(cell is not None for cell in row)]
    if not rows:
        raise ValueError("Sheet '{}' is empty.".format(sheet_name))

    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    data = rows[1:]
    return pd.DataFrame(data, columns=headers)


def _load_definitions(file_obj):
    _seek_start(file_obj)
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message="Workbook contains no default style")
        workbook = load_workbook(file_obj, read_only=False, data_only=True)
    if "Definitions" not in workbook.sheetnames:
        return {}
    definitions = {}
    for row in workbook["Definitions"].iter_rows(values_only=True):
        if row and len(row) >= 2 and row[0] is not None:
            definitions[str(row[0]).strip()] = row[1]
    return definitions


def _parse_onplatform_period(value):
    if not value or "~" not in str(value):
        return pd.NaT, pd.NaT
    start, end = str(value).split("~", 1)
    return pd.to_datetime(start, errors="coerce", dayfirst=True), pd.to_datetime(
        end, errors="coerce", dayfirst=True
    )


def _load_shopee_campaign_type_xlsx(file_obj, source_file):
    definitions = _load_definitions(file_obj)
    df = _worksheet_to_dataframe(file_obj, "By Campaign Type")
    required = {"Campaign Type", "Product Name", "Ads Spend(Local currency)"}
    if not required.issubset(set(df.columns)):
        raise ValueError("Sheet 'By Campaign Type' is not a supported Shopee campaign report.")

    df = df.rename(columns=SHOPEE_CAMPAIGN_TYPE_MAP)
    period_start, period_end = _parse_onplatform_period(definitions.get("Time period"))
    df["platform"] = "Shopee Ads"
    df["account_name"] = definitions.get("Principal name", "")
    df["ad_group_name"] = df["campaign_name"]
    df["bidding_mode"] = df["campaign_name"]
    df["period_start"] = period_start
    df["period_end"] = period_end
    if pd.notna(period_start) and pd.notna(period_end) and period_start == period_end:
        df["date"] = period_start
        df["granularity"] = "daily"
    else:
        df["granularity"] = "period"
    return _finalize(df, source_file)


def load_one_file(file_obj, source_file):
    suffix = Path(source_file).suffix.lower()
    if suffix == ".csv":
        text = _read_text(file_obj)
        if _is_shopee_export(text):
            return _load_shopee_csv(BytesIO(text.encode("utf-8")), source_file)
        df = pd.read_csv(StringIO(text))
        return _load_template_dataframe(df, source_file)

    if suffix == ".xlsx":
        try:
            _seek_start(file_obj)
            with warnings.catch_warnings():
                warnings.filterwarnings("ignore", message="Workbook contains no default style")
                df = pd.read_excel(file_obj, sheet_name="ads_data")
            return _load_template_dataframe(df, source_file)
        except ValueError:
            return _load_shopee_campaign_type_xlsx(file_obj, source_file)

    raise ValueError("Unsupported file type: {}".format(source_file))


def load_ads_files(uploaded_files):
    frames = []
    errors = []
    for uploaded_file in uploaded_files:
        try:
            frames.append(load_one_file(uploaded_file, uploaded_file.name))
        except Exception as exc:
            errors.append("Could not import {}: {}".format(uploaded_file.name, exc))
    if not frames:
        return None, errors
    return pd.concat(frames, ignore_index=True), errors


def load_sample_file(path):
    try:
        return load_one_file(path, str(path.name)), []
    except Exception as exc:
        return None, ["Could not import sample file: {}".format(exc)]
